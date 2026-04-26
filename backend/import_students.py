from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook

from academic_defaults import COURSE_OPTIONS, SEMESTER_OPTIONS, YEAR_LEVEL_OPTIONS
from academic_logic import ensure_schedule_offerings, ensure_section_record, refresh_schedule_student_counts
from app import app
from models import Student, db
from routes.students import (
    build_student_user_account,
    parse_birthday,
    resolve_student_section,
    resolve_student_semester,
    validate_student_payload,
)


STUDENT_FIELD_LIMITS = {
    "student_id": 50,
    "first_name": 100,
    "last_name": 100,
    "middle_name": 100,
    "birthday": 10,
    "email": 120,
    "contact_number": 20,
    "course": 100,
    "year_level": 50,
    "semester": 50,
    "section": 20,
}

STUDENT_COLUMN_ALIASES = {
    "student_id": ["student id", "studentid", "id number", "id"],
    "first_name": ["first name", "firstname", "given name"],
    "last_name": ["last name", "lastname", "surname", "family name"],
    "middle_name": ["middle name", "middlename", "middle initial"],
    "birthday": ["birthday", "birth date", "date of birth", "dob"],
    "email": ["email", "email address"],
    "contact_number": ["contact number", "contact", "phone", "mobile", "contact no"],
    "course": ["course", "program"],
    "year_level": ["year level", "yearlevel", "year", "level"],
    "semester": ["semester", "term"],
    "section": ["section", "sec", "section name"],
}


def sanitize_value(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_header_name(value: str) -> str:
    return "".join(char if char.isalnum() else " " for char in sanitize_value(value).lower()).strip()


def normalize_header_token(value: str) -> str:
    return "".join(char for char in sanitize_value(value).lower() if char.isalnum())


def header_matches_alias(header: str, alias: str) -> bool:
    header_name = normalize_header_name(header)
    alias_name = normalize_header_name(alias)
    if header_name == alias_name:
        return True

    header_token = normalize_header_token(header)
    alias_token = normalize_header_token(alias)
    if not header_token or not alias_token:
        return False
    if header_token == alias_token:
        return True
    if len(alias_token) >= 5 and alias_token in header_token:
        return True
    if len(header_token) >= 5 and header_token in alias_token:
        return True
    return False


def score_header_row(headers: list[str]) -> int:
    matched_fields: set[str] = set()
    for field_key, aliases in STUDENT_COLUMN_ALIASES.items():
        if any(header_matches_alias(header, alias) for alias in aliases for header in headers):
            matched_fields.add(field_key)
    return len(matched_fields)


def convert_matrix_to_objects(matrix: list[list[str]]) -> tuple[list[dict[str, str]], int]:
    meaningful_rows = [row for row in matrix if any(sanitize_value(cell) for cell in row)]
    if not meaningful_rows:
        return [], 0

    best_header_index = -1
    best_score = -1
    max_scan = min(10, len(meaningful_rows))
    for index in range(max_scan):
        headers = [sanitize_value(cell) for cell in meaningful_rows[index]]
        score = score_header_row(headers)
        if score > best_score:
            best_score = score
            best_header_index = index

    if best_header_index < 0:
        return [], 0

    headers = [sanitize_value(cell) for cell in meaningful_rows[best_header_index]]
    data_rows = meaningful_rows[best_header_index + 1 :]
    objects: list[dict[str, str]] = []
    for cells in data_rows:
        row_object: dict[str, str] = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            row_object[header] = sanitize_value(cells[index] if index < len(cells) else "")
        if any(sanitize_value(value) for value in row_object.values()):
            objects.append(row_object)
    return objects, best_score


def resolve_value_from_aliases(row: dict[str, str], aliases: list[str]) -> str:
    for alias in aliases:
        for key, value in row.items():
            if header_matches_alias(key, alias):
                return sanitize_value(value)
    return ""


def clamp_field(value: str, max_length: int) -> str:
    return sanitize_value(value)[:max_length]


def normalize_student_id(value: str) -> str:
    return "".join(char for char in sanitize_value(value) if char.isalnum() or char == "-").upper()


def normalize_name(value: str) -> str:
    return "".join(char for char in sanitize_value(value) if char.isalpha() or char in " .'-")


def normalize_email(value: str) -> str:
    return sanitize_value(value).lower()


def normalize_phone(value: str) -> str:
    return "".join(char for char in sanitize_value(value) if char.isdigit())[:11]


def normalize_course(value: str) -> str:
    raw = sanitize_value(value).upper()
    for course in COURSE_OPTIONS:
        if course.upper() == raw:
            return course
    return raw


def normalize_year_level(value: str) -> str:
    text = sanitize_value(value)
    for option in YEAR_LEVEL_OPTIONS:
        if option.lower() == text.lower():
            return option

    upper = text.upper()
    if "1ST" in upper:
        return "1st Year"
    if "2ND" in upper:
        return "2nd Year"
    if "3RD" in upper:
        return "3rd Year"
    if "4TH" in upper:
        return "4th Year"
    return ""


def normalize_semester(value: str) -> str:
    text = sanitize_value(value)
    for option in SEMESTER_OPTIONS:
        if option.lower() == text.lower():
            return option
    if "2nd" in text.lower():
        return "2nd Semester"
    return "1st Semester" if text else ""


def normalize_birthday(value: str) -> str:
    text = sanitize_value(value)
    if not text:
        return ""
    if len(text) == 10 and text[4] == "-" and text[7] == "-":
        return text

    separators = ["/", "-"]
    for separator in separators:
        parts = text.split(separator)
        if len(parts) == 3 and len(parts[2]) == 4:
            month = parts[0].zfill(2)
            day = parts[1].zfill(2)
            year = parts[2]
            return f"{year}-{month}-{day}"
    return text


def to_student_payload(raw_row: dict[str, str]) -> dict[str, str]:
    section = (
        sanitize_value(resolve_value_from_aliases(raw_row, STUDENT_COLUMN_ALIASES["section"]))
        .replace(" ", "")
        .replace("/", "")
        .upper()
    )
    return {
        "student_id": clamp_field(
            normalize_student_id(resolve_value_from_aliases(raw_row, STUDENT_COLUMN_ALIASES["student_id"])),
            STUDENT_FIELD_LIMITS["student_id"],
        ),
        "first_name": clamp_field(
            normalize_name(resolve_value_from_aliases(raw_row, STUDENT_COLUMN_ALIASES["first_name"])),
            STUDENT_FIELD_LIMITS["first_name"],
        ),
        "last_name": clamp_field(
            normalize_name(resolve_value_from_aliases(raw_row, STUDENT_COLUMN_ALIASES["last_name"])),
            STUDENT_FIELD_LIMITS["last_name"],
        ),
        "middle_name": clamp_field(
            normalize_name(resolve_value_from_aliases(raw_row, STUDENT_COLUMN_ALIASES["middle_name"])),
            STUDENT_FIELD_LIMITS["middle_name"],
        ),
        "birthday": clamp_field(
            normalize_birthday(resolve_value_from_aliases(raw_row, STUDENT_COLUMN_ALIASES["birthday"])),
            STUDENT_FIELD_LIMITS["birthday"],
        ),
        "email": clamp_field(
            normalize_email(resolve_value_from_aliases(raw_row, STUDENT_COLUMN_ALIASES["email"])),
            STUDENT_FIELD_LIMITS["email"],
        ),
        "contact_number": clamp_field(
            normalize_phone(resolve_value_from_aliases(raw_row, STUDENT_COLUMN_ALIASES["contact_number"])),
            STUDENT_FIELD_LIMITS["contact_number"],
        ),
        "course": clamp_field(
            normalize_course(resolve_value_from_aliases(raw_row, STUDENT_COLUMN_ALIASES["course"])),
            STUDENT_FIELD_LIMITS["course"],
        ),
        "year_level": clamp_field(
            normalize_year_level(resolve_value_from_aliases(raw_row, STUDENT_COLUMN_ALIASES["year_level"])),
            STUDENT_FIELD_LIMITS["year_level"],
        ),
        "semester": clamp_field(
            normalize_semester(resolve_value_from_aliases(raw_row, STUDENT_COLUMN_ALIASES["semester"])),
            STUDENT_FIELD_LIMITS["semester"],
        ),
        "section": clamp_field(section, STUDENT_FIELD_LIMITS["section"]),
    }


def load_rows_from_excel(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    matrix = [[cell for cell in row] for row in worksheet.iter_rows(values_only=True)]
    objects, header_score = convert_matrix_to_objects(matrix)
    if not objects:
        return []

    payloads = [to_student_payload(row) for row in objects]
    has_required = any(
        payload["student_id"] and payload["first_name"] and payload["last_name"] and payload["birthday"] and payload["course"] and payload["year_level"]
        for payload in payloads
    )
    if not has_required and header_score < 2:
        raise ValueError(
            "Could not detect student columns. Check headers like Student ID, First Name, Last Name, Birthday, Course, Year Level."
        )
    return payloads


def load_rows_from_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        matrix = [row for row in reader]
    objects, header_score = convert_matrix_to_objects(matrix)
    if not objects:
        return []
    payloads = [to_student_payload(row) for row in objects]
    has_required = any(
        payload["student_id"] and payload["first_name"] and payload["last_name"] and payload["birthday"] and payload["course"] and payload["year_level"]
        for payload in payloads
    )
    if not has_required and header_score < 2:
        raise ValueError(
            "Could not detect student columns. Check headers like Student ID, First Name, Last Name, Birthday, Course, Year Level."
        )
    return payloads


def load_student_payloads(path: Path) -> list[dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return load_rows_from_excel(path)
    if suffix == ".csv":
        return load_rows_from_csv(path)
    raise ValueError("Only .xlsx, .xlsm, and .csv files are supported.")


def import_students(file_path: Path, dry_run: bool = False) -> None:
    payloads = load_student_payloads(file_path)
    if not payloads:
        print("No student rows were detected in the selected file.")
        return

    created_count = 0
    skipped_count = 0
    failed_count = 0

    with app.app_context():
        for index, payload in enumerate(payloads, start=1):
            has_required = all(
                sanitize_value(payload[key])
                for key in ("student_id", "first_name", "last_name", "birthday", "course", "year_level")
            )
            if not has_required:
                skipped_count += 1
                print(f"Row {index}: skipped, missing required fields.")
                continue

            validation_error = validate_student_payload(payload)
            if validation_error:
                skipped_count += 1
                print(f"Row {index}: skipped, {validation_error}.")
                continue

            if dry_run:
                created_count += 1
                print(f"Row {index}: ready to import {payload['student_id']}.")
                continue

            try:
                tenant_id = (payload.get("tenant_id") or "").strip() or None
                birthday = parse_birthday(payload.get("birthday", "").strip())
                semester = resolve_student_semester(payload)
                section = resolve_student_section(payload)
                account = build_student_user_account(payload["student_id"], birthday, tenant_id)

                student = Student(
                    student_id=payload["student_id"],
                    first_name=payload["first_name"],
                    last_name=payload["last_name"],
                    middle_name=payload.get("middle_name") or None,
                    birthday=birthday,
                    email=payload.get("email") or None,
                    contact_number=payload.get("contact_number") or None,
                    course=payload["course"],
                    year_level=payload["year_level"],
                    semester=semester,
                    section=section or None,
                    enrollment_status="Enrolled",
                    tenant_id=tenant_id,
                )

                ensure_section_record(student.course, student.year_level, student.semester, student.section, tenant_id=student.tenant_id)
                db.session.add(account)
                db.session.add(student)
                db.session.commit()

                changed = ensure_schedule_offerings(
                    course=student.course,
                    year_level=student.year_level,
                    semester=student.semester,
                    tenant_id=student.tenant_id,
                )
                if refresh_schedule_student_counts(
                    student.course,
                    student.year_level,
                    student.semester,
                    student.section,
                    tenant_id=student.tenant_id,
                ):
                    changed = True
                if changed:
                    db.session.commit()

                created_count += 1
                print(f"Row {index}: imported {student.student_id}.")
            except Exception as exc:
                db.session.rollback()
                failed_count += 1
                print(f"Row {index}: failed to import {payload.get('student_id') or '(no id)'}: {exc}")

    summary = f"Import finished. Created: {created_count}, Skipped: {skipped_count}, Failed: {failed_count}."
    print(summary)


def main() -> None:
    parser = argparse.ArgumentParser(description="Import students directly from an Excel or CSV file.")
    parser.add_argument("file", help="Path to a .xlsx, .xlsm, or .csv file.")
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Validate and preview rows without writing to MongoDB.",
    )
    args = parser.parse_args()

    file_path = Path(args.file).expanduser()
    if not file_path.exists():
        parser.error(f"File not found: {file_path}")

    import_students(file_path=file_path, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
